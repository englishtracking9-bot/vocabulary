# -*- coding: utf-8 -*-
"""
build_books.py — 匯入 YP 單字書 Excel → data/books.json（YP 專區用）

用法：
    python scripts/build_books.py

會讀取專案根目錄下所有 `YP_levelN_vocabulary.xlsx`（工作表＝Unit1、Unit2…），
產出 data/books.json。只處理 YP 書的「內容」，不碰 data/vocab.json、不碰任何學習進度。

Excel 欄位（每張工作表）：
    序號 | 英文單字 | 詞性 | 中文字 | 英文例句 | 中文翻譯

規則：
- level 來自檔名（YP_level3 → level 3）；unit 來自工作表名（Unit1 → 1）。
- 同一個字連續多列 = 一字多義 → 合併成多個 senses[{pos,zh,example,example_zh}]。
- 片語(phr.)、補充、延伸 等詞性原樣保留；容忍不完整（有些 unit 只有 1-2 列）。
- 每個 entry 有穩定 id（yp{level}-u{unit}-{序號}-{字}），日後補齊重匯只更新內容、
  不影響學習進度（進度存 IndexedDB，依 id/拼字對應）。
- 自帶造句與翻譯直接用，不生成。
"""
import glob
import json
import os
import re
import sys
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "books.json")
HEADER_WORDS = {"英文單字", "英文", "word", "單字"}


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def slug(word):
    s = re.sub(r"[^a-z0-9]+", "-", word.lower()).strip("-")
    return s or "x"


def norm_seq(v):
    s = clean(v)
    return int(s) if s.isdigit() else s


def parse_sheet(ws, level, unit, seen_ids):
    """回傳這張工作表的 entries（已合併一字多義）。"""
    entries = []
    cur = None
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri == 0:
            continue  # 表頭
        seq_raw, word_raw, pos_raw, zh_raw, ex_raw, exzh_raw = (list(row) + [None] * 6)[:6]
        word = clean(word_raw)
        if not word or word in HEADER_WORDS:
            continue
        seq = norm_seq(seq_raw)
        sense = {
            "pos": clean(pos_raw),
            "zh": clean(zh_raw),
            "example": clean(ex_raw),
            "example_zh": clean(exzh_raw),
        }
        # 完全空白的列略過（容忍不完整，但別產生空 sense）
        if not sense["zh"] and not sense["example"] and not sense["pos"]:
            continue
        # 同 (序號, 字) 連續列 → 併入上一個 entry 的 senses
        if cur is not None and cur["_key"] == (seq, word):
            cur["senses"].append(sense)
            continue
        base = f"yp{level}-u{unit}-{seq}-{slug(word)}"
        eid = base
        n = 2
        while eid in seen_ids:
            eid = f"{base}-{n}"
            n += 1
        seen_ids.add(eid)
        cur = {"_key": (seq, word), "id": eid, "seq": seq, "word": word, "senses": [sense]}
        entries.append(cur)
    for e in entries:
        del e["_key"]
    # 依序號排序（數字優先），同序號維持出現順序
    entries.sort(key=lambda e: (0, e["seq"]) if isinstance(e["seq"], int) else (1, str(e["seq"])))
    return entries


def main():
    files = sorted(glob.glob(os.path.join(ROOT, "YP_level*_vocabulary.xlsx")))
    if not files:
        print("找不到 YP_levelN_vocabulary.xlsx（放在專案根目錄）", file=sys.stderr)
        sys.exit(1)

    levels = {}   # level -> {unit -> entries}
    seen_ids = set()
    sources = []
    for path in files:
        fname = os.path.basename(path)
        sources.append(fname)
        m = re.search(r"level(\d+)", fname, re.I)
        if not m:
            print(f"跳過（檔名無 levelN）：{fname}", file=sys.stderr)
            continue
        level = int(m.group(1))
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            um = re.match(r"Unit\s*(\d+)", sheet, re.I)
            if not um:
                continue
            unit = int(um.group(1))
            entries = parse_sheet(wb[sheet], level, unit, seen_ids)
            if entries:
                levels.setdefault(level, {})[unit] = entries
        wb.close()

    # 組成輸出結構
    out_levels = []
    tot_entries = tot_senses = tot_units = 0
    for level in sorted(levels):
        units_out = []
        for unit in sorted(levels[level]):
            ents = levels[level][unit]
            tot_units += 1
            tot_entries += len(ents)
            tot_senses += sum(len(e["senses"]) for e in ents)
            units_out.append({"unit": unit, "count": len(ents), "entries": ents})
        out_levels.append({
            "level": level,
            "unitCount": len(units_out),
            "wordCount": sum(u["count"] for u in units_out),
            "units": units_out,
        })

    data = {
        "book": "YP",
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": sources,
        "levels": out_levels,
        "totals": {"units": tot_units, "entries": tot_entries, "senses": tot_senses},
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f"✅ 寫出 {os.path.relpath(OUT, ROOT)}")
    print(f"   來源：{', '.join(sources)}")
    for lv in out_levels:
        units = ", ".join(f"U{u['unit']}({u['count']})" for u in lv["units"])
        print(f"   Level {lv['level']}：{lv['unitCount']} 單元、{lv['wordCount']} 字 → {units}")
    print(f"   合計：{tot_units} 單元、{tot_entries} 個字、{tot_senses} 個義項")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
